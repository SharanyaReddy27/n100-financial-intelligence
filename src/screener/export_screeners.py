import os
import sys

sys.path.append(os.path.abspath("."))

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from src.screener.engine import load_config, run_screener


OUTPUT_PATH = "output/screener_output.xlsx"

PRESETS = {
    "quality_compounder": "Quality Compounder",
    "value_pick": "Value Pick",
    "growth_accelerator": "Growth Accelerator",
    "dividend_champion": "Dividend Champion",
    "debt_free_blue_chip": "Debt Free Blue Chip",
    "turnaround_watch": "Turnaround Watch",
}

EXPORT_COLUMNS = [
    "company_id",
    "year",
    "broad_sector",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "cfo_quality_ratio",
    "fcf_cagr_5yr",
    "revenue_cagr_3yr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "dividend_payout_ratio_pct",
    "sales",
    "net_profit",
    "market_cap_crore",
    "composite_quality_score",
]


def condition_met(column_name, value, filters):
    if value is None:
        return False

    mapping = {
        "return_on_equity_pct": ("roe_min", ">="),
        "debt_to_equity": ("debt_to_equity_max", "<="),
        "free_cash_flow_cr": ("free_cash_flow_min", ">"),
        "revenue_cagr_5yr": ("revenue_cagr_5yr_min", ">="),
        "revenue_cagr_3yr": ("revenue_cagr_3yr_min", ">="),
        "pat_cagr_5yr": ("pat_cagr_5yr_min", ">="),
        "operating_profit_margin_pct": ("opm_min", ">="),
        "pe_ratio": ("pe_ratio_max", "<="),
        "pb_ratio": ("pb_ratio_max", "<="),
        "dividend_yield_pct": ("dividend_yield_min", ">="),
        "dividend_payout_ratio_pct": ("dividend_payout_max", "<="),
        "interest_coverage": ("icr_min", ">="),
        "market_cap_crore": ("market_cap_min", ">="),
        "net_profit": ("net_profit_min", ">="),
        "eps_cagr_5yr": ("eps_cagr_min", ">="),
        "asset_turnover": ("asset_turnover_min", ">="),
        "sales": ("sales_min", ">="),
    }

    if column_name not in mapping:
        return None

    filter_key, operator = mapping[column_name]

    if filter_key not in filters:
        return None

    threshold = filters[filter_key]

    try:
        if operator == ">=":
            return value >= threshold
        if operator == "<=":
            return value <= threshold
        if operator == ">":
            return value > threshold
    except TypeError:
        return False

    return None


def export_workbook():
    config = load_config()

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:
        for preset_key, sheet_name in PRESETS.items():
            result = run_screener(preset_key)

            available_columns = [
                col for col in EXPORT_COLUMNS
                if col in result.columns
            ]

            export_df = result[available_columns].copy()

            export_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

    format_workbook(config)


def format_workbook(config):
    workbook = load_workbook(OUTPUT_PATH)

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for preset_key, sheet_name in PRESETS.items():
        sheet = workbook[sheet_name]
        filters = config[preset_key]

        headers = {
            cell.value: cell.column
            for cell in sheet[1]
        }

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in range(2, sheet.max_row + 1):
            for column_name, column_index in headers.items():
                cell = sheet.cell(
                    row=row,
                    column=column_index,
                )

                verdict = condition_met(
                    column_name,
                    cell.value,
                    filters,
                )

                if verdict is True:
                    cell.fill = green_fill
                elif verdict is False:
                    cell.fill = red_fill

        for column_index in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(column_index)

            max_length = 0

            for cell in sheet[column_letter]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            sheet.column_dimensions[column_letter].width = min(
                max_length + 2,
                22,
            )

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    workbook.save(OUTPUT_PATH)


if __name__ == "__main__":
    import pandas as pd

    export_workbook()

    print("Screener workbook generated successfully")
    print("Saved to:", OUTPUT_PATH)