from pathlib import Path
import sqlite3

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DB_PATH = Path("db/nifty100.db")
OUTPUT_DIR = Path("output")

SUMMARY_PATH = OUTPUT_DIR / "valuation_summary.xlsx"
FLAGS_PATH = OUTPUT_DIR / "valuation_flags.csv"


def clean_company_id(dataframe):
    """Standardise company ticker values before merging."""
    dataframe = dataframe.copy()

    dataframe["company_id"] = (
        dataframe["company_id"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    return dataframe


def extract_year_number(series):
    """Extract a four-digit year from values such as Mar 2024."""
    return pd.to_numeric(
        series.astype(str).str.extract(r"(\d{4})")[0],
        errors="coerce",
    )


def load_data():
    """Load all valuation inputs from SQLite."""
    if not DB_PATH.exists():
        raise FileNotFoundError(
            f"Database was not found at {DB_PATH.resolve()}"
        )

    with sqlite3.connect(DB_PATH) as connection:
        companies = pd.read_sql_query(
            """
            SELECT
                id AS company_id,
                company_name
            FROM companies
            """,
            connection,
        )

        sectors = pd.read_sql_query(
            """
            SELECT
                company_id,
                broad_sector,
                sub_sector
            FROM sectors
            """,
            connection,
        )

        market_cap = pd.read_sql_query(
            """
            SELECT
                company_id,
                year,
                market_cap_crore,
                enterprise_value_crore,
                pe_ratio,
                pb_ratio,
                ev_ebitda,
                dividend_yield_pct
            FROM market_cap
            """,
            connection,
        )

        ratios = pd.read_sql_query(
            """
            SELECT
                company_id,
                year,
                free_cash_flow_cr
            FROM financial_ratios
            """,
            connection,
        )

    companies = clean_company_id(companies)
    sectors = clean_company_id(sectors)
    market_cap = clean_company_id(market_cap)
    ratios = clean_company_id(ratios)

    return companies, sectors, market_cap, ratios


def get_latest_market_cap(market_cap):
    """Keep the latest available market valuation row per company."""
    data = market_cap.copy()

    data["year_num"] = pd.to_numeric(
        data["year"],
        errors="coerce",
    )

    data = data.dropna(
        subset=["year_num"]
    )

    latest = (
        data.sort_values(
            ["company_id", "year_num"]
        )
        .groupby(
            "company_id",
            as_index=False,
        )
        .tail(1)
    )

    return latest


def get_latest_fcf(ratios):
    """Keep the latest available annual FCF value per company."""
    data = ratios[
        ratios["year"].astype(str).str.upper() != "TTM"
    ].copy()

    data["year_num"] = extract_year_number(
        data["year"]
    )

    data = data.dropna(
        subset=["year_num"]
    )

    latest = (
        data.sort_values(
            ["company_id", "year_num"]
        )
        .groupby(
            "company_id",
            as_index=False,
        )
        .tail(1)
    )

    return latest[
        [
            "company_id",
            "free_cash_flow_cr",
        ]
    ]


def calculate_company_five_year_median_pe(market_cap):
    """
    Calculate each company's median P/E using its latest five
    available market-cap years.
    """
    data = market_cap.copy()

    data["year_num"] = pd.to_numeric(
        data["year"],
        errors="coerce",
    )

    data["pe_ratio"] = pd.to_numeric(
        data["pe_ratio"],
        errors="coerce",
    )

    # Ignore non-positive P/E values because they are not
    # meaningful valuation multiples.
    data.loc[
        data["pe_ratio"] <= 0,
        "pe_ratio",
    ] = np.nan

    latest_five = (
        data.dropna(subset=["year_num"])
        .sort_values(
            ["company_id", "year_num"]
        )
        .groupby(
            "company_id",
            group_keys=False,
        )
        .tail(5)
    )

    median_pe = (
        latest_five.groupby(
            "company_id",
            as_index=False,
        )["pe_ratio"]
        .median()
        .rename(
            columns={
                "pe_ratio": "five_year_median_pe"
            }
        )
    )

    return median_pe


def assign_valuation_flag(pe_ratio, sector_median):
    """Classify valuation relative to the latest sector median P/E."""
    if pd.isna(pe_ratio) or pd.isna(sector_median):
        return "N/A"

    if sector_median <= 0:
        return "N/A"

    if pe_ratio > sector_median * 1.5:
        return "Caution"

    if pe_ratio < sector_median * 0.7:
        return "Discount"

    return "Fair"


def build_valuation_summary():
    companies, sectors, market_cap, ratios = load_data()

    latest_market = get_latest_market_cap(
        market_cap
    )

    latest_fcf = get_latest_fcf(
        ratios
    )

    five_year_median = (
        calculate_company_five_year_median_pe(
            market_cap
        )
    )

    valuation = companies.merge(
        sectors[
            [
                "company_id",
                "broad_sector",
                "sub_sector",
            ]
        ],
        on="company_id",
        how="left",
    )

    valuation = valuation.merge(
        latest_market[
            [
                "company_id",
                "year",
                "market_cap_crore",
                "enterprise_value_crore",
                "pe_ratio",
                "pb_ratio",
                "ev_ebitda",
                "dividend_yield_pct",
            ]
        ],
        on="company_id",
        how="left",
    )

    valuation = valuation.merge(
        latest_fcf,
        on="company_id",
        how="left",
    )

    valuation = valuation.merge(
        five_year_median,
        on="company_id",
        how="left",
    )

    numeric_columns = [
        "market_cap_crore",
        "enterprise_value_crore",
        "pe_ratio",
        "pb_ratio",
        "ev_ebitda",
        "dividend_yield_pct",
        "free_cash_flow_cr",
        "five_year_median_pe",
    ]

    for column in numeric_columns:
        valuation[column] = pd.to_numeric(
            valuation[column],
            errors="coerce",
        )

    # FCF Yield = FCF / Market Capitalisation × 100
    valid_market_cap = (
        valuation["market_cap_crore"] > 0
    )

    valuation["fcf_yield_pct"] = np.where(
        valid_market_cap,
        (
            valuation["free_cash_flow_cr"]
            / valuation["market_cap_crore"]
            * 100
        ),
        np.nan,
    )

    # Do not include negative or zero P/E values in sector medians.
    valuation["valid_pe_for_median"] = (
        valuation["pe_ratio"]
        .where(valuation["pe_ratio"] > 0)
    )

    valuation["sector_median_pe"] = (
        valuation.groupby(
            "broad_sector"
        )["valid_pe_for_median"]
        .transform("median")
    )

    valuation["pe_vs_sector_median_pct"] = np.where(
        valuation["sector_median_pe"] > 0,
        (
            (
                valuation["pe_ratio"]
                - valuation["sector_median_pe"]
            )
            / valuation["sector_median_pe"]
            * 100
        ),
        np.nan,
    )

    valuation["flag"] = valuation.apply(
        lambda row: assign_valuation_flag(
            row["pe_ratio"],
            row["sector_median_pe"],
        ),
        axis=1,
    )

    valuation = valuation.rename(
        columns={
            "broad_sector": "sector",
            "pe_ratio": "pe_ratio",
            "pb_ratio": "pb_ratio",
            "ev_ebitda": "ev_ebitda",
            "five_year_median_pe": "5yr_median_PE",
        }
    )

    output_columns = [
        "company_id",
        "company_name",
        "sector",
        "sub_sector",
        "year",
        "market_cap_crore",
        "enterprise_value_crore",
        "pe_ratio",
        "pb_ratio",
        "ev_ebitda",
        "dividend_yield_pct",
        "free_cash_flow_cr",
        "fcf_yield_pct",
        "5yr_median_PE",
        "sector_median_pe",
        "pe_vs_sector_median_pct",
        "flag",
    ]

    valuation = valuation[
        output_columns
    ].copy()

    valuation = valuation.sort_values(
        [
            "sector",
            "company_id",
        ],
        na_position="last",
    ).reset_index(drop=True)

    return valuation


def save_to_sqlite(valuation):
    """
    Save valuation output for the Streamlit get_valuation()
    database helper.
    """
    with sqlite3.connect(DB_PATH) as connection:
        valuation.to_sql(
            "valuation",
            connection,
            if_exists="replace",
            index=False,
        )


def format_summary_workbook():
    workbook = load_workbook(
        SUMMARY_PATH
    )

    worksheet = workbook["Valuation Summary"]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    caution_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    discount_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    fair_fill = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    header_map = {
        cell.value: cell.column
        for cell in worksheet[1]
    }

    flag_column = header_map.get(
        "flag"
    )

    if flag_column:
        for row_number in range(
            2,
            worksheet.max_row + 1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=flag_column,
            )

            if cell.value == "Caution":
                cell.fill = caution_fill
            elif cell.value == "Discount":
                cell.fill = discount_fill
            elif cell.value == "Fair":
                cell.fill = fair_fill

    percentage_columns = {
        "fcf_yield_pct",
        "pe_vs_sector_median_pct",
        "dividend_yield_pct",
    }

    for header, column_number in header_map.items():
        if header in percentage_columns:
            for row_number in range(
                2,
                worksheet.max_row + 1,
            ):
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).number_format = "0.00"

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        max_length = 0

        for cell in worksheet[
            column_letter
        ]:
            value = (
                ""
                if cell.value is None
                else str(cell.value)
            )

            max_length = max(
                max_length,
                len(value),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            28,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    workbook.save(
        SUMMARY_PATH
    )


def export_outputs(valuation):
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    valuation.to_excel(
        SUMMARY_PATH,
        sheet_name="Valuation Summary",
        index=False,
        engine="openpyxl",
    )

    flagged = valuation[
        valuation["flag"].isin(
            ["Caution", "Discount"]
        )
    ].copy()

    flagged.to_csv(
        FLAGS_PATH,
        index=False,
    )

    save_to_sqlite(
        valuation
    )

    format_summary_workbook()

    return flagged


def validate_outputs(valuation, flagged):
    print(
        "Valuation module completed successfully"
    )
    print(
        "Summary rows:",
        len(valuation),
    )
    print(
        "Unique companies:",
        valuation["company_id"].nunique(),
    )
    print(
        "Caution flags:",
        int(
            valuation["flag"]
            .eq("Caution")
            .sum()
        ),
    )
    print(
        "Discount flags:",
        int(
            valuation["flag"]
            .eq("Discount")
            .sum()
        ),
    )
    print(
        "Fair flags:",
        int(
            valuation["flag"]
            .eq("Fair")
            .sum()
        ),
    )
    print(
        "Flag-file rows:",
        len(flagged),
    )
    print(
        "Saved:",
        SUMMARY_PATH,
    )
    print(
        "Saved:",
        FLAGS_PATH,
    )
    print(
        valuation.head()
    )


def main():
    valuation = build_valuation_summary()

    flagged = export_outputs(
        valuation
    )

    validate_outputs(
        valuation,
        flagged,
    )


if __name__ == "__main__":
    main()