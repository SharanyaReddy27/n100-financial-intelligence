import os
import sqlite3
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

sys.path.append(os.path.abspath("."))

DB_PATH = "db/nifty100.db"
OUTPUT_FILE = "output/peer_comparison.xlsx"


def load_data():
    conn = sqlite3.connect(DB_PATH)

    peer_groups = pd.read_sql_query(
        "SELECT * FROM peer_groups",
        conn,
    )

    financial_ratios = pd.read_sql_query(
        "SELECT * FROM financial_ratios",
        conn,
    )

    peer_percentiles = pd.read_sql_query(
        "SELECT * FROM peer_percentiles",
        conn,
    )

    conn.close()

    for df in [peer_groups, financial_ratios, peer_percentiles]:
        df["company_id"] = (
            df["company_id"]
            .astype(str)
            .str.strip()
            .str.upper()
        )

    financial_ratios = financial_ratios[
        financial_ratios["year"].str.upper() != "TTM"
    ].copy()

    financial_ratios["year_num"] = pd.to_numeric(
        financial_ratios["year"].str.extract(r"(\d{4})")[0],
        errors="coerce",
    )

    financial_ratios = (
        financial_ratios
        .sort_values(["company_id", "year_num"])
        .groupby("company_id", as_index=False)
        .tail(1)
    )

    return peer_groups, financial_ratios, peer_percentiles


def build_peer_dataframe(
    peer_group_name,
    peer_groups,
    financial_ratios,
    peer_percentiles,
):
    companies = peer_groups[
        peer_groups["peer_group_name"] == peer_group_name
    ]

    report = companies.merge(
        financial_ratios,
        on="company_id",
        how="left",
    )

    pivot = (
        peer_percentiles[
            peer_percentiles["peer_group_name"] == peer_group_name
        ]
        .pivot(
            index="company_id",
            columns="metric",
            values="percentile_rank",
        )
        .reset_index()
    )

    rename = {}

    for col in pivot.columns:
        if col != "company_id":
            rename[col] = col + "_pct_rank"

    pivot.rename(columns=rename, inplace=True)

    report = report.merge(
        pivot,
        on="company_id",
        how="left",
    )

    return report
from openpyxl.styles import Alignment


GREEN = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

YELLOW = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)

RED = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC",
)

HEADER = PatternFill(
    fill_type="solid",
    fgColor="D9EAD3",
)

BENCHMARK = PatternFill(
    fill_type="solid",
    fgColor="FFD966",
)


def format_sheet(sheet):
    """
    Format worksheet after writing data.
    """

    headers = {}

    for cell in sheet[1]:
        headers[cell.value] = cell.column
        cell.font = Font(bold=True)
        cell.fill = HEADER
        cell.alignment = Alignment(horizontal="center")

    # -----------------------------
    # Highlight benchmark company
    # -----------------------------
    if "is_benchmark" in headers:
        benchmark_col = headers["is_benchmark"]

        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, benchmark_col).value:
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row, col).fill = BENCHMARK

    # -----------------------------
    # Colour percentile columns
    # -----------------------------
    for name, col in headers.items():

        if not str(name).endswith("_pct_rank"):
            continue

        for row in range(2, sheet.max_row + 1):

            cell = sheet.cell(row, col)

            if cell.value is None:
                continue

            value = float(cell.value)

            if value >= 0.75:
                cell.fill = GREEN

            elif value <= 0.25:
                cell.fill = RED

            else:
                cell.fill = YELLOW

    # -----------------------------
    # Auto column width
    # -----------------------------
    for column in sheet.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None else 0
            for cell in column
        )

        sheet.column_dimensions[
            column[0].column_letter
        ].width = min(length + 2, 24)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def add_median_row(sheet):
    """
    Add median row at bottom of worksheet.
    """

    last = sheet.max_row + 1

    sheet.cell(last, 1).value = "Median"

    headers = {
        cell.value: cell.column
        for cell in sheet[1]
    }

    for name, col in headers.items():

        if str(name).endswith("_pct_rank"):

            letter = sheet.cell(1, col).column_letter

            sheet.cell(last, col).value = (
                f"=MEDIAN({letter}2:{letter}{last-1})"
            )

    sheet.cell(last, 1).font = Font(bold=True)
def generate_report():
    peer_groups, financial_ratios, peer_percentiles = load_data()

    unique_groups = sorted(
        peer_groups["peer_group_name"].dropna().unique()
    )

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        for group_name in unique_groups:

            report = build_peer_dataframe(
                group_name,
                peer_groups,
                financial_ratios,
                peer_percentiles,
            )

            report = report.sort_values(
                "composite_quality_score",
                ascending=False,
            )

            report.to_excel(
                writer,
                sheet_name=group_name[:31],
                index=False,
            )

    workbook = load_workbook(OUTPUT_FILE)

    for sheet in workbook.worksheets:
        format_sheet(sheet)
        add_median_row(sheet)

    workbook.save(OUTPUT_FILE)

    print("=" * 60)
    print("Peer comparison workbook generated successfully")
    print("Output:", OUTPUT_FILE)
    print("Worksheets:", len(workbook.sheetnames))
    print("=" * 60)

    for sheet in workbook.sheetnames:
        print(sheet)


if __name__ == "__main__":
    generate_report()